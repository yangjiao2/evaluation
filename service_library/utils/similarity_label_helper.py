import pandas as pd
import pandas as pd
from collections import Counter
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from itertools import combinations
from bert_score import score

SEPARATOR = ":"
import pandas as pd

### ---
# loading file
### ---


def load_and_sort_excel(file_name: str) -> pd.DataFrame:
    """Loads and sorts the Excel file by the first column."""
    df = pd.read_excel(file_name)
    df_sorted = df.sort_values(by=df.columns[0])  # Sort by the first column (assumed to be 'ID')
    return df_sorted


def extract_columns(df: pd.DataFrame, columns: list, file_index: int = None,
                    column_suffix_separator: str = "_") -> pd.DataFrame:
    """
    Extracts columns from the DataFrame and labels them based on the file index.

    :param df: Input DataFrame.
    :param columns: List of column names to extract.
    :param file_index: Index to append to the column name.
    :param column_suffix_separator: String to use as the separator between column names and file index.
    :return: Extracted DataFrame with renamed columns.
    """
    extracted_columns = {}
    for col in columns:
        if col in df.columns:
            col_label = f"{col}{column_suffix_separator}{file_index}" if file_index is not None else col
            extracted_columns[col_label] = df[col]
        else:
            print(f"Error: Column '{col}' not found in file {file_index}")
    return pd.DataFrame(extracted_columns)


def align_comparison_columns(dfs: list, comparison_columns: list, column_suffix_separator: str = "_") -> pd.DataFrame:
    """
    Aligns comparison columns from all files side by side, ensuring they are grouped for easy comparison.

    :param dfs: List of DataFrames (from each file).
    :param comparison_columns: List of comparison columns to align.
    :param column_suffix_separator: Separator used for naming columns.
    :return: A DataFrame with aligned comparison columns.
    """
    aligned_columns = {}

    for col in comparison_columns:
        aligned_col_group = []
        for i, df in enumerate(dfs, start=1):
            col_label = f"{col}{column_suffix_separator}{i}"
            print ("column_label", col_label)
            if col_label in df.columns:
                aligned_col_group.append(df[col_label])
            else:
                print(f"Error: Column '{col_label}' not found in file {i}")

        # Combine the aligned columns for this specific comparison column
        aligned_columns[col] = pd.concat(aligned_col_group, axis=1)

    # Concatenate all aligned columns for the final DataFrame
    return pd.concat(aligned_columns.values(), axis=1)


def process_files(file_names: list, additional_columns: list, comparison_columns: list,
                  column_suffix_separator: str = "_"):
    """
    Processes multiple Excel files, extracts specified columns, and merges them into one DataFrame.

    :param file_names: List of Excel file paths.
    :param additional_columns: List of columns to extract from the first file.
    :param comparison_columns: List of columns to extract from the rest of the files.
    :param column_suffix_separator: Separator used for naming columns from different files.
    :return: A merged DataFrame with sorted and extracted data.
    """
    # Step 1: Load and sort the first file (additional_columns)
    df_first = load_and_sort_excel(file_names[0])

    # Step 2: Extract additional_columns from the first file
    additional_data = extract_columns(df_first, additional_columns)

    # Step 3: Loop through the rest of the files and extract comparison_columns
    comparison_dataframes = []
    for i, file_name in enumerate(file_names[1:], start=1):
        df = load_and_sort_excel(file_name)
        comparison_data = extract_columns(df, comparison_columns, i, column_suffix_separator)
        comparison_dataframes.append(comparison_data)

    # Step 4: Align comparison columns side by side
    df_aligned_comparison = align_comparison_columns(comparison_dataframes, comparison_columns, column_suffix_separator)

    # Step 5: Merge additional_data with aligned comparison columns
    df_final = pd.concat([additional_data, df_aligned_comparison], axis=1)

    return df_final

# Optionally, save the result to Excel
# df_final.to_excel('merged_output.xlsx', index=False)

### ---
# outlier label
### ---


def find_max_count(strings):
    # Count occurrences of each string in the list
    string_counts = Counter(strings)

    # Find the maximum count
    max_count = max(string_counts.values())

    # Find all strings that have the maximum count
    most_common_strings = [string for string, count in string_counts.items() if count == max_count]

    return most_common_strings, max_count


def label_outliers(df: pd.DataFrame, columns: list, threshold: float) -> pd.DataFrame:
    """
    Adds a new column '<Column Name> Outlier' to the DataFrame. Marks columns as outliers if any pairwise comparison
    score is below the threshold. Adds light red formatting to cells below the threshold.

    :param df: The input DataFrame.
    :param columns: A list of column names used in pairwise comparisons.
    :param threshold: The threshold value for outlier detection.
    :return: The DataFrame with the new '<Column Name> Outlier' columns.
    """
    # Iterate over the rows and check for outliers
    for index, row in df.iterrows():
        outlier_columns = set()
        outlier_column_info = []

        for col in columns:
            if row[col] < threshold:
                outlier_columns.add(col)
                for col2 in col.split(SEPARATOR):
                    outlier_column_info.append(col2)

        # Update the 'Outlier Response' column for the row
        df.at[index, 'Outlier Response'] = ', '.join(
            [column_name for column_name in find_max_count(outlier_column_info)[0]])

        # Add individual '<Column Name> Outlier' columns
        for col in outlier_columns:
            df.at[index, f'{col.split(SEPARATOR)[0]} Outlier'] = True

    return df


def read_excel_first_sheet(file_path: str) -> pd.DataFrame:
    """
    Reads the first sheet of an Excel file into a DataFrame.
    """
    df = pd.read_excel(file_path, sheet_name=0)  # 'sheet_name=0' reads the first sheet
    return df


def add_pairwise_bert_f1_to_dataframe(df: pd.DataFrame, columns: list) -> pd.DataFrame:
    """
    Adds new columns with BERT F1 scores for each pairwise combination of the specified columns.
    The new columns are named '<col1>:<col2>'.

    :param df: The input DataFrame.
    :param columns: A list of four column names to compare pairwise.
    :return: The DataFrame with new BERT F1 score columns added.
    """
    # Generate all pairwise combinations of the columns
    pairs = list(combinations(columns, 2))
    print("columns pairs:", pairs)
    # For each pair of columns, calculate BERT F1 scores and add as new columns
    for col1, col2 in pairs:
        gt_list = df[col1].tolist()
        pt_list = df[col2].tolist()

        # Calculate BERT F1 scores for the pair
        _, _, f1 = score([str(g) for g in gt_list], [str(p) for p in pt_list], lang="en",
                         model_type="bert-base-uncased", batch_size=64)

        # Add the F1 scores as a new column with the name "<col1>:<col2>"
        score_col = f"{col1}{SEPARATOR}{col2}"
        df[score_col] = f1.tolist()

    return df


def save_outliers_to_excel(input_excel: str, columns: list, threshold: float, output_excel: str = None):
    """
    Reads the first sheet of an Excel file, adds pairwise BERT F1 scores, labels outliers,
    creates a second sheet with only outliers, and saves the updated DataFrame back to Excel.

    :param input_excel: Path to the input Excel file.
    :param columns: A list of column names to compare pairwise.
    :param threshold: The threshold value for outlier detection.
    :param output_excel: Path to the output Excel file. If None, overwrites the input file.
    """
    # Step 1: Read the first sheet of the Excel file
    df = load_and_sort_excel(input_excel)

    # Step 2: Add pairwise BERT F1 scores to the DataFrame
    df = add_pairwise_bert_f1_to_dataframe(df, columns)

    # Generate BERT F1 score columns for outlier detection
    score_columns = [f"{col1}{SEPARATOR}{col2}" for col1, col2 in combinations(columns, 2)]

    # Step 3: Label outliers and add new '<Column Name> Outlier' columns
    if threshold:
        df = label_outliers(df, score_columns, threshold)

    print(f"Updated Excel file saved to {output_excel}")
    return df


def apply_color_to_outliers(output_excel: str, df: pd.DataFrame, columns: list, threshold: float):
    """
    Applies light red color formatting to cells below the threshold in the outlier columns.

    :param output_excel: The Excel file to apply formatting to.
    :param df: The DataFrame containing the data.
    :param columns: The list of column names used for outlier detection.
    :param threshold: The threshold value for outlier detection.
    """
    # Load the Excel file
    wb = load_workbook(output_excel)
    ws = wb.active

    # Define a light red fill for cells below the threshold
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Apply the color to the relevant cells
    for col in columns:
        for row in range(2, len(df) + 2):  # Excel rows start at 1, skip the header
            cell_value = ws[f'{col}{row}'].value
            if isinstance(cell_value, (int, float)) and cell_value < threshold:
                ws[f'{col}{row}'].fill = red_fill

    # Save the workbook with the new formatting
    wb.save(output_excel)


# Example usage:
# save_outliers_to_excel('input.xlsx', ['col1', 'col2', 'col3', 'col4'], 0.7)


def load_and_sort_excel(input_excel: str, output_excel: str = None):
    # Step 1: Load the Excel file into a DataFrame
    df = pd.read_excel(input_excel)

    # Step 2: Sort the DataFrame by the first column (ID)
    # Assumption: The first column is named 'ID'
    df_sorted = df.sort_values(by=df.columns[0])  # Sort by the first column

    # Step 3: Save the sorted DataFrame back to Excel (optional)
    if output_excel:
        df_sorted.to_excel(output_excel, index=False)
        print(f"Sorted Excel file saved to {output_excel}")

    return df_sorted

# Example usage:
# sorted_df = load_and_sort_excel('input.xlsx', 'sorted_output.xlsx')
#     # Step 4: Save the updated DataFrame to Excel with light red color formatting
#     if output_excel is None:
#         output_excel = input_excel
#
#     with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
#         # Save the main DataFrame with '<Column Name> Outlier' columns
#         df.to_excel(writer, index=False, sheet_name='Sheet1')
#
#     # Apply light red fill to outlier cells
#     apply_color_to_outliers(output_excel, df, score_columns, threshold)
